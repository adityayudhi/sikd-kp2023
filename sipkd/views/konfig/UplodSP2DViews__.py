from django.shortcuts import render,redirect
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.db import connection,connections
from sipkd.config import *
from django.urls import reverse
from django.contrib import messages
from django.db import IntegrityError
from support.support_sipkd import *

import openpyxl
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# import numpy as np

def uplod_sp2dgaji(request):
	tahun = tahun_log(request)
	bulan = 0

	with connections[tahun_log(request)].cursor() as cursor:
		cursor.execute("SELECT DISTINCT(extract (month from tglsp2d)) as bulan \
			from log.temp_gaji where tahun=%s",[tahun])
		datax = dictfetchall(cursor)

	for x in datax:
		bulan = x['bulan']

	data ={'bulan':int(bulan)}
	return render(request, 'konfig/uplod_sp2dgaji.html',data)

def uplod_sp2dgaji_list(request):
	tahun = tahun_log(request)
	data  = request.POST
	bulan = int(data.get('bln'))

	with connections[tahun_log(request)].cursor() as cursor:
		cursor.execute("SELECT row_number() over(ORDER BY kodeurusan, kodesuburusan, kodeorganisasi ASC) as nomor,\
			kodeurusan||'.'||lpad(kodesuburusan::text,2,'0')||'.'||lpad(kodeorganisasi::text,2,'0') as kode, namaskpd,\
			tglgaji,tglposting,tglsp2d,tglspm,nomorsp2d,nomorspm,keterangan,norekpenerima,namapenerima,bankpenerima,npwp,triwulan,\
			gapok,t_keluarga,t_jabatan,t_fungsional,t_umum,t_beras,t_pph,t_pembulatan,t_bpjs,t_jkk,t_jkm,t_tkd,t_tdt,\
			p_iwpn8,p_iwpn2,p_askes,p_pph,p_bulog,p_taperum,p_sewarumah,p_jkk,p_jkm,p_lainnya,\
			(gapok+t_keluarga+t_jabatan+t_fungsional+t_umum+t_beras+t_pph+t_pembulatan+t_bpjs+t_jkk+t_jkm+t_tkd+t_tdt) as jml_sp2d,\
			(p_iwpn8+p_iwpn2+p_askes+p_pph+p_bulog+p_taperum+p_sewarumah+p_jkk+p_jkm+p_lainnya) as jml_potongan \
			FROM log.temp_gaji WHERE tahun = %s and extract (month from tglsp2d) = %s \
			ORDER BY kodeurusan, kodesuburusan, kodeorganisasi",[tahun,bulan])
		hasil = ArrayFormater(dictfetchall(cursor))

	data = {'list_uplod':hasil}
	return render(request, 'konfig/tabel/uplod_sp2dgaji_list.html', data)

def uplod_sp2dgaji_temp(request):
	isSimpan = 0
	pesan 	 = ""
	data 	 = request.POST
	piles 	 = request.FILES
	tahun 	 = tahun_log(request)

	xlsx_file = piles.get('uploadfile')
	wb_obj = openpyxl.load_workbook(xlsx_file)
	sheet = wb_obj.active

	# book = Workbook()
	# sheet2 = book.active
	with connections[tahun_log(request)].cursor() as cursor:
		cursor.execute("DELETE FROM log.temp_gaji where tahun = %s",[tahun])

	array_to_db = []
	array_to_db_ = []
	for x in range(1,sheet.max_row):
		array_to_db_ = []
		for column in range(sheet.max_column):
			columnletter=get_column_letter(column+1)

			if(sheet["%s%s"%(columnletter,x+1)].value != None):
				try:
					data = sheet["%s%s"%(columnletter,x+1)].value.strip()
				except Exception as e:
					data = sheet["%s%s"%(columnletter,x+1)].value
					
				array_to_db_.append(data)
			else:
				pass

		# PERINTAH INSERT GOES HERE
		# array_to_db.append(array_to_db_)

		with connections[tahun_log(request)].cursor() as cursor:
			try:
				cursor.execute("INSERT INTO log.temp_gaji \
					(tahun, kdjnstrans, jenisgaji, kddati1, kddati2, tglgaji, kdskpd, \
					kodeurusan, kodesuburusan, kodeorganisasi, namaskpd, nomorsp2d, \
					tglsp2d, nomorspm, tglspm, tglposting, keterangan, gapok, t_keluarga, t_jabatan, t_fungsional, \
					t_umum, t_beras, t_pph, t_pembulatan, t_bpjs, t_jkk, t_jkm, t_tkd, t_tdt, \
					p_iwpn8, p_iwpn2, p_askes, p_pph, p_bulog, p_taperum, p_sewarumah, p_jkk, p_jkm, p_lainnya, \
					norekpenerima, namapenerima, npwp, bankpenerima, triwulan)\
					VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,\
					%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",array_to_db_)
				messages.success(request, "Unggah data SP2D Gaji, Sukses !")
			except Exception as e:
				pesan = 'Unggah data SP2D Gaji, GAGAL !'
				messages.error(request, "Unggah data SP2D Gaji, GAGAL !")
				break	
	return redirect('sipkd:uplod_sp2dgaji')


def uplod_sp2dgaji_dump(request):
	isSimpan = 0
	pesan 	 = ""
	data 	 = request.POST
	tahun 	 = tahun_log(request)
	bulan    = int(data.get('bln'))

	with connections[tahun_log(request)].cursor() as cursor:
		try:
			cursor.execute("SELECT log.fc_dump_gaji_perbulan(%s,%s)",[tahun,bulan])
			isSimpan = 1
			pesan 	 = 'Unggah data SP2D Gaji BERHASIL !'
		except IntegrityError as e:
			isSimpan = 0
			pesan 	 = 'Unggah data SP2D Gaji GAGAL, silahkan unggah ulang dan cek Nomor SP2D !'

	with connections[tahun_log(request)].cursor() as cursor:
		cursor.execute("DELETE FROM log.temp_gaji where tahun = %s",[tahun])

	hasil = {'pesan':pesan, 'issimpan':isSimpan}
	return JsonResponse(hasil)

def uplod_sp2dgaji_hapus(request):
	tahun = tahun_log(request)

	with connections[tahun_log(request)].cursor() as cursor:
		try:
			cursor.execute("DELETE FROM log.temp_gaji where tahun = %s",[tahun])
			isSimpan = 1
			pesan 	 = 'Tabel Gaji BERHASIL dikosongkan !'
		except IntegrityError as e:
			isSimpan = 0
			pesan 	 = 'Pengosongan tabel Gaji GAGAL!'

	hasil = {'pesan':pesan, 'issimpan':isSimpan}
	return JsonResponse(hasil)

## ==============================================================================================
## UPLOD DATA EXCEL DARI APLIKASI SIPD kemendagri
## ==============================================================================================
def uplod_excelsipd(request):
	tahun = tahun_log(request)

	jns_dataExcel = [
            {'id': 'SP2D', 'uraian': 'Excel SP2D SIPD'},
            {'id': 'PENDAPATAN', 'uraian': 'Excel Pendapatan SIPD'},
    	]

	data = {'list_jenis':jns_dataExcel}

	return render(request, 'konfig/uplod_excelsipd.html',data)

def uplod_excelsipd_list(request):
	tahun = tahun_log(request)
	data  = request.POST
	hasil = ''

	jns_dt = data.get('jns_dt').upper()
	kdskpd = data.get('kdskpd')

	if data.get('skpd') != "":
		skpd = data.get('skpd').split(".")
	else:
		skpd = '0.0.0.0'.split('.')


	if jns_dt == 'SP2D':
		with connections[tahun_log(request)].cursor() as cursor:
			cursor.execute("SELECT row_number() over(ORDER BY tglsp2d ASC) as nomor,\
				kodeskpd,namaskpd,nosp2d,tglsp2d,kodeakun,namaakun,bruto::numeric,potongan::numeric \
				FROM temporary_sipd.registersp2d  WHERE kodeskpd = %s \
				ORDER BY nomor",[kdskpd])
			hasil = ArrayFormater(dictfetchall(cursor))

	elif jns_dt == 'PENDAPATAN':
		with connections[tahun_log(request)].cursor() as cursor:
			cursor.execute("SELECT row_number() over(ORDER BY tanggal ASC) as nomor,\
				kodeskpd,skpd,nodokumen,tanggal,koderekening,namarekening,nilai::numeric \
				FROM temporary_sipd.registersts  WHERE kodeskpd = %s \
				ORDER BY nomor ",[kdskpd])
			hasil = ArrayFormater(dictfetchall(cursor))


	data = {'jns_dt':jns_dt, 'list_uplod':hasil}
	return render(request, 'konfig/tabel/uplod_excelsipd_list.html', data)

def uplod_excelsipd_temp(request):
	isSimpan = 0
	pesan 	 = ""
	data 	 = request.POST
	piles 	 = request.FILES
	tahun 	 = tahun_log(request)

	eskapede = data.get('organisasi')
	jns_dt = data.get('jenis_data').upper()
	kodeskpd = data.get('kodeskpd')

	xlsx_file = piles.get('uploadfile')
	wb_obj = openpyxl.load_workbook(xlsx_file)
	sheet = wb_obj.active

	## book = Workbook()
	## sheet2 = book.active
	if jns_dt == 'SP2D':
		with connections[tahun_log(request)].cursor() as cursor:
			cursor.execute("DELETE FROM temporary_sipd.registersp2d WHERE kodeskpd = %s",[kodeskpd])
	elif jns_dt == 'PENDAPATAN':
		with connections[tahun_log(request)].cursor() as cursor:
			cursor.execute("DELETE FROM temporary_sipd.registersts")

	array_to_db = []
	array_to_db_ = []
	for x in range(1,sheet.max_row):
		array_to_db_ = []
		for column in range(sheet.max_column):
			columnletter=get_column_letter(column+1)

			# if(sheet["%s%s"%(columnletter,x+1)].value != None):
			try:
				data = sheet["%s%s"%(columnletter,x+1)].value.strip()
			except Exception as e:
				data = sheet["%s%s"%(columnletter,x+1)].value
				
			array_to_db_.append(data)
			# else:
			# 	pass

		if jns_dt == 'SP2D':
			with connections[tahun_log(request)].cursor() as cursor:
				try:
					cursor.execute("INSERT INTO temporary_sipd.registersp2d \
						(kodeskpd,namaskpd,nospp,tglspp,nospm,tglspm,nosp2d,tglsp2d,jenis,uraian,\
						penerima,rekening,bank,npwp,kodeprogram,namaprogram,kodekegiatan,\
						namakegiatan,kodesubkegiatan,namasubkegiatan,kodeakun,namaakun,\
						bruto,potongan,netto,tglposting,no_lpj) \
						VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s,\
							%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
							%s, %s, %s, %s, %s, %s, %s)", array_to_db_)

					pesan = 'Unggah data Excel, Sukses !'
					messages.success(request, "Unggah data Excel, Sukses !")
				except Exception as e:
					pesan = 'Unggah data Excel, GAGAL !'
					messages.error(request, "Unggah data Excel, GAGAL !")
					break	

		elif jns_dt == 'PENDAPATAN':
			with connections[tahun_log(request)].cursor() as cursor:
				try:
					cursor.execute("INSERT INTO temporary_sipd.registersts \
						(no, nodokumen, tanggal, koderekening, namarekening,\
						nilai, keterangan, kodeskpd, skpd, kodeunit, unit) \
						VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", array_to_db_)

					pesan = 'Unggah data Excel, Sukses !'
					messages.success(request, "Unggah data Excel, Sukses !")
				except Exception as e:
					pesan = 'Unggah data Excel, GAGAL !'
					messages.error(request, "Unggah data Excel, GAGAL !")
					break

	hasil = {'pesan':pesan, 'issimpan':isSimpan}
	return JsonResponse(hasil)


def uplod_excelsipd_dump(request):
	isSimpan = 0
	pesan 	 = ""
	data 	 = request.POST
	tahun 	 = tahun_log(request)
	kdskpd   = data.get('kdskpd')
	jns_dt = data.get('jns_dt').upper() 

	if jns_dt == 'SP2D':
		with connections[tahun_log(request)].cursor() as cursor:
			# try:
			cursor.execute("SELECT temporary_sipd.fc_push_sp2d(%s,%s)",[tahun,kdskpd])
			isSimpan = 1
			pesan 	 = 'Unggah data SP2D BERHASIL !'
			# except IntegrityError as e:
			# 	isSimpan = 0
			# 	pesan 	 = 'Unggah data SP2D GAGAL, silahkan unggah ulang dan cek Nomor SP2D !'

	elif jns_dt == 'PENDAPATAN':
		with connections[tahun_log(request)].cursor() as cursor:
			# try:
			cursor.execute("SELECT temporary_sipd.fc_push_kasda_sts(%s)",[tahun])
			isSimpan = 1
			pesan 	 = 'Unggah data SP2D BERHASIL !'
			# except IntegrityError as e:
			# 	isSimpan = 0
			# 	pesan 	 = 'Unggah data SP2D GAGAL, silahkan unggah ulang dan cek Nomor SP2D !'
	


	hasil = {'pesan':pesan, 'issimpan':isSimpan}
	return JsonResponse(hasil)


def uplod_excelsipd_hapus(request):
	isSimpan = 0
	pesan 	 = ""
	tahun 	 = tahun_log(request)
	kdskpd   = request.POST.get('kdskpd')
	jns_dt = request.POST.get('jns_dt').upper()

	if jns_dt == 'SP2D':
		with connections[tahun_log(request)].cursor() as cursor:
			try:
				cursor.execute("DELETE FROM temporary_sipd.registersp2d where kodeskpd = %s",[kdskpd])
				isSimpan = 1
				pesan 	 = 'Tabel BERHASIL dikosongkan !'
			except IntegrityError as e:
				isSimpan = 0
				pesan 	 = 'Pengosongan tabel GAGAL!'

	elif jns_dt == 'PENDAPATAN':
		with connections[tahun_log(request)].cursor() as cursor:
			try:
				cursor.execute("DELETE FROM temporary_sipd.registersts")
				isSimpan = 1
				pesan 	 = 'Tabel BERHASIL dikosongkan !'
			except IntegrityError as e:
				isSimpan = 0
				pesan 	 = 'Pengosongan tabel GAGAL!'
	

	hasil = {'pesan':pesan, 'issimpan':isSimpan}
	return JsonResponse(hasil)